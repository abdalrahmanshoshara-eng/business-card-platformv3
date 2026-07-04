from __future__ import annotations

import hashlib
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from cards.models import BusinessCard
from cards.services.card_data import INVESTMENT_TYPE_CHOICES, combine_bilingual, prepare_card_data
from cards.services.normalization import normalize_text
from cards.services.website_enrichment import INVESTMENT_TYPE_KEYWORDS

EMPTY_MARKERS = {'', '-', '—', '_', 'لا يوجد', 'لايوجد', 'n/a', 'na', 'none', 'null'}
EMAIL_RE = re.compile(r'[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+')
ARABIC_RE = re.compile(r'[\u0600-\u06FF]')


@dataclass
class ImportResult:
    created: int = 0
    updated: int = 0
    skipped_empty: int = 0
    skipped_duplicate: int = 0
    needs_review: int = 0
    errors: int = 0


def clean_cell(value: Any) -> str:
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        text = str(int(value))
    else:
        text = str(value)
    text = text.replace('\u00a0', ' ')
    text = re.sub(r'\s+', ' ', text).strip()
    if text.lower() in EMPTY_MARKERS:
        return ''
    return text


def split_values(value: str) -> list[str]:
    value = clean_cell(value)
    if not value:
        return []
    parts = re.split(r'[\n;,،|]+|\s+/\s+|/+', value)
    return [clean_cell(part) for part in parts if clean_cell(part)]


def extract_emails(value: str) -> tuple[list[str], list[str]]:
    value = clean_cell(value).replace(' ', '')
    if not value:
        return [], []
    matches = [match.group(0).strip('.').lower() for match in EMAIL_RE.finditer(value)]
    matches = list(dict.fromkeys(matches))
    invalid = [email for email in matches if '.' not in email.split('@')[-1]]
    return matches, invalid


def extract_phones(value: Any) -> list[str]:
    raw = clean_cell(value)
    if not raw:
        return []
    parts = split_values(raw)
    if not parts:
        parts = [raw]
    return parts


def normalize_url(value: str) -> str:
    value = clean_cell(value).replace(' ', '')
    if not value:
        return ''
    # Keep only the first URL if a spreadsheet cell contains alternatives.
    value = re.split(r'[\n;,،|]+', value)[0].strip()
    return value


def looks_arabic(value: str) -> bool:
    return bool(ARABIC_RE.search(value or ''))


def infer_investment_type_offline(*parts: str) -> tuple[str, str]:
    combined = ' '.join(part for part in parts if part).lower()
    for investment_type, keywords in INVESTMENT_TYPE_KEYWORDS:
        if any(str(keyword).lower() in combined for keyword in keywords):
            return investment_type, ''

    # Common short labels used in the legacy Excel file.
    compact = normalize_text(combined)
    manual_map = [
        ('غذائية', 'مؤسسة الصناعات الغذائية'),
        ('اغذية', 'مؤسسة الصناعات الغذائية'),
        ('هندسية', 'المؤسسة العامة للصناعات الهندسية'),
        ('كهرباء', 'المؤسسة العامة للصناعات الهندسية'),
        ('ميكانيك', 'المؤسسة العامة للصناعات الهندسية'),
        ('نسيج', 'المؤسسة العامة للصناعات النسيجية'),
        ('البسة', 'المؤسسة العامة للصناعات النسيجية'),
        ('كيميائية', 'المؤسسة العامة للصناعات الكيميائية'),
        ('دوائية', 'المؤسسة العامة للصناعات الكيميائية'),
        ('اسمنت', 'المؤسسة العامة للصناعات الكيميائية'),
        ('تعدين', 'هيئة ادارة المعادن النبيلة وهيئة المواصفات و المقاييس'),
        ('معادن', 'هيئة ادارة المعادن النبيلة وهيئة المواصفات و المقاييس'),
        ('مختبر', 'مركز الاختبارات و الابحاث'),
        ('اختبارات', 'مركز الاختبارات و الابحاث'),
        ('ابحاث', 'مركز الاختبارات و الابحاث'),
        ('تدريب', 'مديرية الاشراف على التاهيل الفني'),
        ('تعليم', 'مديرية الاشراف على التاهيل الفني'),
    ]
    for keyword, official in manual_map:
        if keyword in compact:
            return official, ''

    return 'غير ذلك', ''


def build_raw_text(source_sheet: str, source_row: int, columns: dict[str, str]) -> str:
    lines = [f'Excel source: {source_sheet} row {source_row}']
    for label, value in columns.items():
        if value:
            lines.append(f'{label}: {value}')
    return '\n'.join(lines)


def row_to_payload(sheet_name: str, row_number: int, row: tuple[Any, ...]) -> tuple[dict | None, list[str]]:
    values = [clean_cell(item) for item in row[:13]]
    if len(values) < 13:
        values.extend([''] * (13 - len(values)))

    excel_seq, company, person_ar, person_en, job_title, phone_raw, email_raw, country, city, raw_investment_type, activity, website, notes = values

    sequence_number = None
    if excel_seq:
        try:
            sequence_number = int(float(excel_seq))
        except (TypeError, ValueError):
            sequence_number = None

    meaningful = [company, person_ar, person_en, job_title, phone_raw, email_raw, website, activity, notes]
    if not any(meaningful):
        return None, []

    review_notes: list[str] = []
    emails, invalid_emails = extract_emails(email_raw)
    if invalid_emails:
        review_notes.append('بريد إلكتروني يحتاج تدقيق: ' + ', '.join(invalid_emails))

    phones = extract_phones(phone_raw)
    address = ' - '.join(item for item in [country, city] if item)

    activity_parts = []
    if activity:
        activity_parts.append(activity)
    if notes and notes not in activity_parts:
        activity_parts.append(notes)
    company_activity = ' - '.join(activity_parts)

    investment_type = raw_investment_type if raw_investment_type in INVESTMENT_TYPE_CHOICES else ''
    investment_type_other = ''
    if not investment_type:
        investment_type, investment_type_other = infer_investment_type_offline(company, company_activity, raw_investment_type)
        if raw_investment_type and raw_investment_type not in {'غير ذلك'} and raw_investment_type not in INVESTMENT_TYPE_CHOICES:
            investment_type_other = investment_type_other or raw_investment_type

    company_ar = company if looks_arabic(company) else ''
    company_en = company if company and not looks_arabic(company) else ''

    columns = {
        'م': excel_seq,
        'اسم الشركة': company,
        'اسم الشخص عربي': person_ar,
        'اسم الشخص إنجليزي': person_en,
        'الصفة': job_title,
        'رقم الهاتف': phone_raw,
        'البريد الإلكتروني': email_raw,
        'الدولة': country,
        'المدينة': city,
        'نوع الاستثمار': raw_investment_type,
        'طبيعة الاستثمار': activity,
        'الموقع الإلكتروني': website,
        'ملاحظات': notes,
    }

    payload = {
        'person_name_ar': person_ar if looks_arabic(person_ar) else '',
        'person_name_en': person_en or (person_ar if person_ar and not looks_arabic(person_ar) else ''),
        'job_title_ar': job_title if looks_arabic(job_title) else '',
        'job_title_en': job_title if job_title and not looks_arabic(job_title) else '',
        'company_name_ar': company_ar,
        'company_name_en': company_en,
        'mobile_numbers': phones,
        'emails': emails,
        'website': normalize_url(website),
        'address': address,
        'company_activity': company_activity,
        'investment_type': investment_type,
        'investment_type_other': investment_type_other,
        'raw_text': build_raw_text(sheet_name, row_number, columns),
        'confidence': 1.0,
        'needs_review': False,
        'review_notes': ' | '.join(review_notes),
        'website_visit_note': '',
        'status': 'reviewed',
    }
    if sequence_number is not None:
        payload['sequence_number'] = sequence_number

    payload['person_name'] = combine_bilingual(payload['person_name_ar'], payload['person_name_en'])
    payload['job_title'] = combine_bilingual(payload['job_title_ar'], payload['job_title_en'])
    payload['company_name'] = combine_bilingual(payload['company_name_ar'], payload['company_name_en'], company)

    if not payload['company_name'] or not payload['person_name'] or not payload['mobile_numbers']:
        payload['needs_review'] = True
        payload['status'] = 'needs_review'
        review_notes.append('حقول أساسية ناقصة من ملف Excel')
    if invalid_emails:
        payload['needs_review'] = True
        payload['status'] = 'needs_review'
    if review_notes:
        payload['review_notes'] = ' | '.join(dict.fromkeys(review_notes))

    return payload, review_notes


class Command(BaseCommand):
    help = 'Import the legacy business-card Excel workbook into the BusinessCard database.'

    def add_arguments(self, parser):
        parser.add_argument('excel_path', type=str, help='Path to the .xlsx file inside the backend container or local environment.')
        parser.add_argument('--clear', action='store_true', help='Delete all existing cards before import. Use only when the database is disposable.')
        parser.add_argument('--update-existing', action='store_true', help='Update existing rows when duplicate_hash already exists instead of skipping them.')
        parser.add_argument('--dry-run', action='store_true', help='Parse the Excel file and print a summary without writing to the database.')
        parser.add_argument('--max-rows', type=int, default=0, help='Optional limit for testing.')
        parser.add_argument('--sheet', type=str, default='', help='Import only one worksheet by exact name. Recommended for the legacy cards file.')
        parser.add_argument('--preserve-sequence', action='store_true', help='Use the Excel م column as BusinessCard.sequence_number and keep duplicate_hash unique per Excel ID.')

    def handle(self, *args, **options):
        excel_path = Path(options['excel_path'])
        if not excel_path.exists():
            raise CommandError(f'Excel file not found: {excel_path}')

        result = ImportResult()
        parsed_payloads: list[dict] = []

        workbook = load_workbook(excel_path, data_only=True)
        selected_sheet = options.get('sheet') or ''
        if selected_sheet:
            if selected_sheet not in workbook.sheetnames:
                raise CommandError(f"Worksheet not found: {selected_sheet}. Available: {', '.join(workbook.sheetnames)}")
            worksheets = [workbook[selected_sheet]]
        else:
            worksheets = workbook.worksheets

        for worksheet in worksheets:
            for row_number, row in enumerate(worksheet.iter_rows(min_row=2, max_col=13, values_only=True), start=2):
                if options['max_rows'] and len(parsed_payloads) >= options['max_rows']:
                    break
                payload, _notes = row_to_payload(worksheet.title, row_number, row)
                if payload is None:
                    result.skipped_empty += 1
                    continue
                prepared = prepare_card_data(payload, infer_missing_investment=False)
                if options.get('preserve_sequence') and payload.get('sequence_number'):
                    prepared['sequence_number'] = payload['sequence_number']
                    prepared['duplicate_hash'] = hashlib.sha256(
                        f"legacy_excel_sequence:{payload['sequence_number']}".encode('utf-8')
                    ).hexdigest()
                parsed_payloads.append(prepared)
                if prepared.get('needs_review'):
                    result.needs_review += 1
            if options['max_rows'] and len(parsed_payloads) >= options['max_rows']:
                break

        if options['dry_run']:
            self.stdout.write(self.style.WARNING('DRY RUN - no database writes were performed.'))
            self._print_summary(result, parsed_payloads)
            return

        with transaction.atomic():
            if options['clear']:
                BusinessCard.objects.all().delete()

            seen_hashes: set[str] = set()
            for payload in parsed_payloads:
                duplicate_hash = payload['duplicate_hash']
                if duplicate_hash in seen_hashes:
                    result.skipped_duplicate += 1
                    continue
                seen_hashes.add(duplicate_hash)

                existing = None
                if options.get('preserve_sequence') and payload.get('sequence_number'):
                    existing = BusinessCard.objects.filter(sequence_number=payload['sequence_number']).first()
                if not existing:
                    existing = BusinessCard.objects.filter(duplicate_hash=duplicate_hash).first()
                if existing:
                    if not options['update_existing']:
                        result.skipped_duplicate += 1
                        continue
                    for field, value in payload.items():
                        if field not in {'sequence_number'}:
                            setattr(existing, field, value)
                    existing.save()
                    result.updated += 1
                    continue

                BusinessCard.objects.create(**payload)
                result.created += 1

        self.stdout.write(self.style.SUCCESS('Excel import finished.'))
        self._print_summary(result, parsed_payloads)

    def _print_summary(self, result: ImportResult, parsed_payloads: list[dict]) -> None:
        self.stdout.write(f'Parsed rows: {len(parsed_payloads)}')
        self.stdout.write(f'Created: {result.created}')
        self.stdout.write(f'Updated: {result.updated}')
        self.stdout.write(f'Skipped empty rows: {result.skipped_empty}')
        self.stdout.write(f'Skipped duplicates: {result.skipped_duplicate}')
        self.stdout.write(f'Rows needing review: {result.needs_review}')
