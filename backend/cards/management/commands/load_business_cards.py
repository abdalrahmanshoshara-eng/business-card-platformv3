"""Load cards from an export-format .xlsx (the columns produced by export-xlsx).

Usage:
    python manage.py load_business_cards                       # imports/business-cards.xlsx
    python manage.py load_business_cards --flush               # delete existing cards first
    python manage.py load_business_cards --owner admin         # owner (default: earliest superuser)
    python manage.py load_business_cards --file /path/to.xlsx

Every imported card is owned by the given admin (owner comes from the server,
never from the sheet). Country is derived from the address/phone.
"""
from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path

from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandError
from django.db import IntegrityError, transaction
from django.utils import timezone
from openpyxl import load_workbook

from cards.models import BusinessCard
from cards.services.card_data import merge_missing_card_data, prepare_card_data
from cards.services.duplicates import salt_duplicate_hash
from cards.services.merge import resequence_cards

User = get_user_model()

# Column header -> position is resolved by name so column order is flexible.
COL = {
    'seq': '#', 'person': 'اسم الشخص', 'title_ar': 'المنصب (عربي)', 'title_en': 'المنصب (إنجليزي)',
    'company': 'الشركة', 'phones': 'الموبايلات', 'emails': 'الإيميلات', 'website': 'الموقع',
    'address': 'العنوان', 'activity': 'نشاط الشركة', 'inv_type': 'نوع الاستثمار',
    'inv_other': 'تفاصيل الاستثمار', 'needs_review': 'يحتاج مراجعة', 'created': 'تاريخ الإضافة',
}


def _s(v) -> str:
    if v is None:
        return ''
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _split(v) -> list[str]:
    text = _s(v)
    if not text:
        return []
    return [part.strip() for part in re.split(r'[|,\n]+', text) if part.strip()]


def _parse_dt(v):
    if isinstance(v, datetime):
        dt = v
    else:
        text = _s(v)
        if not text:
            return None
        dt = None
        for fmt in ('%Y-%m-%d %H:%M', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d'):
            try:
                dt = datetime.strptime(text, fmt)
                break
            except ValueError:
                continue
        if dt is None:
            return None
    if timezone.is_naive(dt):
        dt = timezone.make_aware(dt, timezone.get_current_timezone())
    return dt


class Command(BaseCommand):
    help = 'Import business cards from an export-format .xlsx and assign them to an admin.'

    def add_arguments(self, parser):
        parser.add_argument('--file', default='imports/business-cards.xlsx')
        parser.add_argument('--owner', help='Username of the owner (default: earliest superuser).')
        parser.add_argument('--flush', action='store_true', help='Delete all existing cards first.')

    def handle(self, *args, **opts):
        path = Path(opts['file'])
        if not path.is_absolute():
            from django.conf import settings
            path = Path(settings.BASE_DIR) / path
        if not path.exists():
            raise CommandError(f'الملف غير موجود: {path}')

        # Exact country per sequence number (built by comparing the approved
        # sheet with the bilingual country sheet). Overrides the derived value.
        country_map = {}
        cmap_path = Path(__file__).resolve().parent.parent.parent / 'data' / 'country_by_sequence.json'
        if cmap_path.exists():
            country_map = json.loads(cmap_path.read_text(encoding='utf-8'))

        if opts.get('owner'):
            try:
                owner = User.objects.get(username=opts['owner'])
            except User.DoesNotExist:
                raise CommandError(f'المستخدم "{opts["owner"]}" غير موجود.')
        else:
            owner = User.objects.filter(is_superuser=True).order_by('date_joined', 'id').first()
            if owner is None:
                raise CommandError('لا يوجد superuser. أنشئ حساب مشرف أولاً أو مرّر --owner.')
        if not (owner.is_staff or owner.is_superuser):
            raise CommandError(f'المستخدم "{owner.username}" ليس مشرفاً.')

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise CommandError('الملف فارغ.')
        header = [(_s(h)) for h in rows[0]]
        try:
            idx = {key: header.index(name) for key, name in COL.items()}
        except ValueError as exc:
            raise CommandError(f'عمود مفقود في الملف: {exc}')

        created = skipped = merged = 0
        with transaction.atomic():
            if opts['flush']:
                deleted = BusinessCard.objects.count()
                BusinessCard.objects.all().delete()
                self.stdout.write(self.style.WARNING(f'تم حذف {deleted} كرت قبل الاستيراد.'))

            for raw in rows[1:]:
                def cell(key):
                    return raw[idx[key]] if idx[key] < len(raw) else None

                person = _s(cell('person'))
                company = _s(cell('company'))
                if not (person or company or _s(cell('emails')) or _s(cell('phones'))):
                    skipped += 1
                    continue

                needs_review = _s(cell('needs_review')) in {'نعم', 'yes', 'true', '1'}
                data = {
                    'person_name': person,
                    'job_title_ar': _s(cell('title_ar')),
                    'job_title_en': _s(cell('title_en')),
                    'company_name': company,
                    'mobile_numbers': _split(cell('phones')),
                    'emails': _split(cell('emails')),
                    'website': _s(cell('website')),
                    'address': _s(cell('address')),
                    'company_activity': _s(cell('activity')),
                    'investment_type': _s(cell('inv_type')),
                    'investment_type_other': _s(cell('inv_other')),
                    'raw_text': '',
                    'needs_review': needs_review,
                    'status': 'needs_review' if needs_review else 'new',
                }
                seq = _s(cell('seq'))
                if seq in country_map:
                    data['country'] = country_map[seq]
                prepared = prepare_card_data(data, infer_missing_investment=False)

                extra = {'owner': owner}
                if seq.isdigit() and not BusinessCard.objects.filter(sequence_number=int(seq)).exists():
                    extra['sequence_number'] = int(seq)
                created_at = _parse_dt(cell('created'))
                if created_at:
                    extra['created_at'] = created_at

                try:
                    with transaction.atomic():
                        BusinessCard.objects.create(**prepared, **extra)
                    created += 1
                except IntegrityError:
                    # Genuine duplicate contact for this owner: merge into the
                    # existing card (union contacts, fill blanks) so nothing is
                    # lost and no duplicate row is created.
                    existing = BusinessCard.objects.filter(
                        owner=owner, duplicate_hash=prepared['duplicate_hash']
                    ).first()
                    if existing is not None:
                        merge_missing_card_data(existing, prepared)
                        merged += 1
                    else:
                        prepared = salt_duplicate_hash(prepared)
                        with transaction.atomic():
                            BusinessCard.objects.create(**prepared, **extra)
                        created += 1

        resequence_cards(BusinessCard.objects.all())

        self.stdout.write(self.style.SUCCESS(
            f'تم استيراد {created} كرت، ودمج {merged} تكرار، وتخطّي {skipped}، '
            f'وإعادة الترقيم تسلسلياً، وإسنادها إلى "{owner.username}".'
        ))
