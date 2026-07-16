from __future__ import annotations

import os
import logging
import re
import tempfile
import time

from django.db import DatabaseError, IntegrityError
from django.db.models import Q
from django.http import FileResponse, Http404, HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from rest_framework import pagination, status, viewsets
from rest_framework.decorators import action, api_view
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from .models import BusinessCard
from .serializers import BusinessCardSerializer
from .services.extractor import ExtractionError, extract_business_card
from .services.image_processing import preprocess_image
from .services.normalization import duplicate_reason
from .services.card_data import INVESTMENT_TYPE_CHOICES, merge_missing_card_data, merge_missing_card_images, prepare_card_data
from .services.natural_search import apply_natural_search, derive_category, parse_natural_query
from .services.access import cards_for_user
from .services.duplicates import (
    find_duplicate_candidates,
    find_existing_duplicate,
    salt_duplicate_hash,
)
from .services.security import excel_safe, validate_image_upload

logger = logging.getLogger(__name__)


def _truthy(value) -> bool:
    if isinstance(value, bool):
        return value
    return str(value or '').strip().lower() in {'1', 'true', 'yes', 'y', 'on'}


def _validation_error_response(exc) -> Response:
    detail = getattr(exc, 'detail', None)
    if isinstance(detail, dict):
        messages = []
        for field, errors in detail.items():
            if isinstance(errors, (list, tuple)):
                joined = ' '.join(str(error) for error in errors)
            else:
                joined = str(errors)
            messages.append(f'{field}: {joined}')
        message = '؛ '.join(messages) or 'بيانات غير صالحة.'
    else:
        message = str(detail or 'بيانات غير صالحة.')
    return Response({'detail': message, 'error_type': 'validation_error', 'errors': detail}, status=status.HTTP_400_BAD_REQUEST)


def _prepare_data(data: dict) -> dict:
    return prepare_card_data(data)


class BusinessCardPagination(pagination.PageNumberPagination):
    page_size = 20
    page_size_query_param = 'page_size'
    max_page_size = 100


class BusinessCardViewSet(viewsets.ModelViewSet):
    queryset = BusinessCard.objects.all().order_by('-sequence_number', '-id')
    serializer_class = BusinessCardSerializer
    pagination_class = BusinessCardPagination

    def _scoped(self):
        """Ownership-scoped base queryset for the current request user."""
        return cards_for_user(self.request.user)

    def _find_existing_duplicate(self, data: dict):
        candidates = find_duplicate_candidates(data, self._scoped(), limit=1)
        if not candidates or candidates[0]['score'] < 80:
            return None
        return self._scoped().filter(id=candidates[0]['id']).first()

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if not serializer.is_valid():
            logger.warning('manual_card_validation_failed errors=%s', serializer.errors)
            return Response(
                {
                    'detail': 'تعذر حفظ الكرت بسبب بيانات غير صالحة. يرجى مراجعة الحقول.',
                    'error_type': 'validation_error',
                    'errors': serializer.errors,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        validated = dict(serializer.validated_data)
        front = request.FILES.get('front') or request.FILES.get('front_image') or validated.pop('front_image', None)
        back = request.FILES.get('back') or request.FILES.get('back_image') or validated.pop('back_image', None)
        try:
            validate_image_upload(front)
            validate_image_upload(back)
        except Exception as exc:
            return _validation_error_response(exc)
        data = _prepare_data(validated)
        confirm_duplicate = _truthy(request.data.get('confirm_duplicate') or request.data.get('force_create'))

        candidates = find_duplicate_candidates(data, self._scoped())
        strong_candidates = [candidate for candidate in candidates if candidate['score'] >= 80]
        if strong_candidates and not confirm_duplicate:
            logger.info('manual_card_duplicate_detected candidates=%s', len(strong_candidates))
            return Response(
                {
                    'detail': 'يوجد كرت مشابه بنفس البريد أو رقم الهاتف أو نفس بيانات الشركة/الشخص.',
                    'error_type': 'duplicate_card',
                    'duplicate_conflict': True,
                    'duplicate_candidates': strong_candidates,
                },
                status=status.HTTP_409_CONFLICT,
            )

        if strong_candidates and confirm_duplicate:
            data = salt_duplicate_hash(data)
            logger.info('manual_card_duplicate_confirmed candidates=%s', len(strong_candidates))

        save_kwargs = dict(data)
        save_kwargs['owner'] = request.user  # owner comes from the session, never the client
        if front is not None:
            save_kwargs['front_image'] = front
        if back is not None:
            save_kwargs['back_image'] = back

        try:
            card = serializer.save(**save_kwargs)
        except IntegrityError:
            logger.exception('manual_card_integrity_error')
            candidates = find_duplicate_candidates(data, self._scoped())
            if candidates and not confirm_duplicate:
                return Response(
                    {
                        'detail': 'يوجد كرت مشابه بنفس البريد أو رقم الهاتف.',
                        'error_type': 'duplicate_card',
                        'duplicate_conflict': True,
                        'duplicate_candidates': candidates,
                    },
                    status=status.HTTP_409_CONFLICT,
                )
            return Response(
                {'detail': 'تعذر حفظ الكرت بسبب تعارض في قاعدة البيانات. يرجى المحاولة مرة أخرى.', 'error_type': 'database_integrity_error'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
        except DatabaseError:
            logger.exception('manual_card_database_error')
            return Response(
                {'detail': 'تعذر حفظ بيانات الكرت في قاعدة البيانات.', 'error_type': 'database_save_error'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        logger.info('manual_card_created card_id=%s duplicate_candidates=%s', card.id, len(candidates))
        headers = self.get_success_headers(BusinessCardSerializer(card, context={'request': request}).data)
        return Response(
            {
                'duplicate': False,
                'saved': True,
                'card': BusinessCardSerializer(card, context={'request': request}).data,
                'message': f'تم حفظ الكرت كسجل رقم {card.sequence_number}',
            },
            status=status.HTTP_201_CREATED,
            headers=headers,
        )

    def get_queryset(self):
        qs = cards_for_user(self.request.user)
        q = self.request.query_params.get('q', '').strip()
        if q:
            parsed = parse_natural_query(q)
            qs = apply_natural_search(qs, parsed)

        company = self.request.query_params.get('company', '').strip()
        if company:
            qs = qs.filter(
                Q(company_name__icontains=company) |
                Q(company_name_ar__icontains=company) |
                Q(company_name_en__icontains=company)
            )

        activity = self.request.query_params.get('activity', '').strip()
        if activity:
            qs = qs.filter(company_activity__icontains=activity)

        country = self.request.query_params.get('country', '').strip()
        if country:
            qs = qs.filter(country__iexact=country)

        # Admin-only: view a specific user's cards (used by the admin UI).
        owner_param = self.request.query_params.get('owner', '').strip()
        if owner_param.isdigit() and (self.request.user.is_staff or self.request.user.is_superuser):
            qs = qs.filter(owner_id=int(owner_param))

        created_from = self.request.query_params.get('created_from', '').strip()
        if created_from:
            qs = qs.filter(created_at__date__gte=created_from)
        created_to = self.request.query_params.get('created_to', '').strip()
        if created_to:
            qs = qs.filter(created_at__date__lte=created_to)

        # Exact-match filter used by the "stats by category" panel, since
        # those buckets are derived (see derive_category) and don't always
        # equal the raw company_activity text.
        category = self.request.query_params.get('category', '').strip()
        if category:
            matching_ids = [
                card_id for card_id, activity_value in qs.values_list('id', 'company_activity')
                if derive_category(activity_value) == category
            ]
            qs = qs.filter(id__in=matching_ids)

        investment_type = self.request.query_params.get('investment_type', '').strip()
        if investment_type:
            qs = qs.filter(Q(investment_type__icontains=investment_type) | Q(investment_type_other__icontains=investment_type))

        needs_review = self.request.query_params.get('needs_review')
        if needs_review in {'true', 'false'}:
            qs = qs.filter(needs_review=(needs_review == 'true'))

        status_value = self.request.query_params.get('status', '').strip()
        if status_value:
            qs = qs.filter(status=status_value)

        sort_value = self.request.query_params.get('sort', 'newest').strip().lower()
        if sort_value == 'oldest':
            return qs.order_by('sequence_number', 'id')
        return qs.order_by('-sequence_number', '-id')

    def perform_create(self, serializer):
        validated = dict(serializer.validated_data)
        validated.pop('front_image', None)
        validated.pop('back_image', None)
        data = _prepare_data(validated)
        # Owner is always the session user; never accepted from the client.
        serializer.save(owner=self.request.user, **data)

    def perform_update(self, serializer):
        current = {
            field.name: getattr(serializer.instance, field.name)
            for field in BusinessCard._meta.fields
            if field.name not in {'id', 'owner', 'sequence_number', 'created_at', 'updated_at', 'front_image', 'back_image'}
        }
        validated = dict(serializer.validated_data)
        front_direct = validated.pop('front_image', None)
        back_direct = validated.pop('back_image', None)
        current.update(validated)
        data = prepare_card_data(current, touched_fields=set(validated))
        # The recomputed duplicate_hash may collide with another card owned by
        # the same user (e.g. genuine duplicates kept as separate rows). Salt it
        # so the edit succeeds without violating the per-owner unique constraint.
        new_hash = data.get('duplicate_hash')
        if new_hash and BusinessCard.objects.filter(
            owner_id=serializer.instance.owner_id, duplicate_hash=new_hash
        ).exclude(pk=serializer.instance.pk).exists():
            data['duplicate_hash'] = salt_duplicate_hash(data)['duplicate_hash']
        # Allow uploaded front/back files to replace existing images when editing
        front = self.request.FILES.get('front') or front_direct
        back = self.request.FILES.get('back') or back_direct
        save_kwargs = dict(data)
        if front is not None:
            save_kwargs['front_image'] = front
        if back is not None:
            save_kwargs['back_image'] = back
        serializer.save(**save_kwargs)

    @action(detail=False, methods=['post'], url_path='extract')
    def extract(self, request):
        endpoint_started = time.perf_counter()
        front = request.FILES.get('front')
        back = request.FILES.get('back')
        if not front:
            return Response({'detail': 'يجب رفع صورة الوجه الأمامي.'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            validate_image_upload(front)
            validate_image_upload(back)
        except Exception as exc:
            return _validation_error_response(exc)

        with tempfile.TemporaryDirectory() as tmpdir:
            preprocessing_started = time.perf_counter()
            try:
                front_path = os.path.join(tmpdir, front.name or 'front.jpg')
                with open(front_path, 'wb') as file_handle:
                    for chunk in front.chunks():
                        file_handle.write(chunk)
                front_processed = preprocess_image(front_path)

                back_processed = None
                if back:
                    back_path = os.path.join(tmpdir, back.name or 'back.jpg')
                    with open(back_path, 'wb') as file_handle:
                        for chunk in back.chunks():
                            file_handle.write(chunk)
                    back_processed = preprocess_image(back_path)
            except Exception:
                logger.exception('card_preprocessing_failed')
                return Response(
                    {'detail': 'تعذرت معالجة صورة الكرت. يرجى تجربة صورة أوضح أو أصغر.', 'error_type': 'preprocessing_error'},
                    status=status.HTTP_422_UNPROCESSABLE_ENTITY,
                )

            logger.info(
                'card_preprocessing_done elapsed_ms=%d front_bytes=%s back_bytes=%s',
                int((time.perf_counter() - preprocessing_started) * 1000),
                os.path.getsize(front_processed) if front_processed else None,
                os.path.getsize(back_processed) if back_processed else None,
            )

            extraction_started = time.perf_counter()
            try:
                extracted = extract_business_card(front_processed, back_processed).model_dump()
            except ExtractionError as exc:
                logger.warning(
                    'card_extraction_failed category=%s status=%s recoverable=%s elapsed_ms=%d',
                    exc.category,
                    exc.status_code,
                    getattr(exc, 'recoverable', True),
                    int((time.perf_counter() - extraction_started) * 1000),
                )
                return Response(
                    {
                        'detail': str(exc),
                        'error_type': exc.category,
                        'recoverable': getattr(exc, 'recoverable', True),
                    },
                    status=exc.status_code,
                )
            except Exception as exc:
                msg = str(exc)
                logger.exception('card_extraction_unhandled elapsed_ms=%d', int((time.perf_counter() - extraction_started) * 1000))
                if '429' in msg or 'RESOURCE_EXHAUSTED' in msg:
                    retry_seconds = None
                    m = re.search(r'retry[^\d]*(\d+(?:\.\d+)?)\s*s', msg, re.I)
                    if m:
                        retry_seconds = int(float(m.group(1))) + 1
                    detail = 'تم تجاوز حد Gemini أو معدل الطلبات.'
                    if retry_seconds:
                        detail += f' يرجى الانتظار {retry_seconds} ثانية ثم المحاولة مجدداً.'
                    return Response({'detail': detail, 'error_type': 'gemini_quota_exceeded'}, status=status.HTTP_429_TOO_MANY_REQUESTS)
                return Response({'detail': 'فشل استخراج البيانات من Gemini. يرجى المحاولة مجدداً.', 'error_type': 'unknown_extraction_error'}, status=status.HTTP_502_BAD_GATEWAY)

        prepared = _prepare_data(extracted)
        # Duplicate lookup is scoped to the current user's own cards.
        existing = find_existing_duplicate(prepared, self._scoped())

        if existing:
            updated_fields = [
                *merge_missing_card_data(existing, prepared),
                *merge_missing_card_images(existing, front, back),
            ]
            logger.info(
                'card_duplicate_handled elapsed_ms=%d existing_id=%s updated_fields=%s',
                int((time.perf_counter() - endpoint_started) * 1000),
                existing.id,
                updated_fields,
            )
            return Response({
                'duplicate': True,
                'saved': bool(updated_fields),
                'updated': bool(updated_fields),
                'updated_fields': updated_fields,
                'reason': duplicate_reason(prepared, existing),
                'existing_card': BusinessCardSerializer(existing, context={'request': request}).data,
                'extracted_data': extracted,
            }, status=status.HTTP_200_OK)

        front.seek(0)
        if back:
            back.seek(0)

        try:
            card = BusinessCard.objects.create(
                **prepared,
                owner=request.user,
                status='needs_review' if prepared.get('needs_review') else 'new',
                front_image=front,
                back_image=back if back else None,
            )
        except IntegrityError:
            logger.exception('card_save_integrity_failed')
            existing = self._find_existing_duplicate(prepared)
            if existing:
                updated_fields = [
                    *merge_missing_card_data(existing, prepared),
                    *merge_missing_card_images(existing, front, back),
                ]
                return Response({
                    'duplicate': True,
                    'saved': bool(updated_fields),
                    'updated': bool(updated_fields),
                    'updated_fields': updated_fields,
                    'reason': duplicate_reason(prepared, existing),
                    'existing_card': BusinessCardSerializer(existing, context={'request': request}).data,
                    'extracted_data': extracted,
                }, status=status.HTTP_200_OK)
            return Response(
                {'detail': 'تعذر حفظ الكرت بسبب تعارض في قاعدة البيانات. يرجى المحاولة مرة أخرى.', 'error_type': 'database_integrity_error'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
        except DatabaseError:
            logger.exception('card_save_failed')
            return Response(
                {'detail': 'تعذر حفظ بيانات الكرت في قاعدة البيانات.', 'error_type': 'database_save_error'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        logger.info('card_extract_endpoint_done elapsed_ms=%d card_id=%s', int((time.perf_counter() - endpoint_started) * 1000), card.id)
        return Response({
            'duplicate': False,
            'saved': True,
            'card': BusinessCardSerializer(card, context={'request': request}).data,
            'message': f'تم حفظ الكرت كسجل رقم {card.sequence_number}',
        }, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['get'], url_path='stats')
    def stats(self, request):
        qs = self._scoped()
        return Response({
            'total': qs.count(),
            'needs_review': qs.filter(needs_review=True).count(),
            'companies': qs.exclude(company_name='').values('company_name').distinct().count(),
            'with_email': qs.exclude(emails=[]).count(),
            'with_phone': qs.exclude(mobile_numbers=[]).count(),
        })

    @action(detail=False, methods=['get'], url_path='stats-by-category')
    def stats_by_category(self, request):
        """Count companies/cards grouped by specialty (activity or investment type)."""
        field = request.query_params.get('field', 'activity').strip().lower()
        qs = self._scoped().only(
            'company_activity', 'investment_type', 'investment_type_other', 'company_name'
        )

        # Count cards (rows) per bucket so the chip number matches exactly what
        # the table shows when that category is used as a filter.
        buckets: dict[str, int] = {}
        for card in qs:
            if field == 'investment_type':
                category = (card.investment_type or '').strip() or 'غير محدد'
            else:
                category = derive_category(card.company_activity)
            buckets[category] = buckets.get(category, 0) + 1

        result = [
            {'category': category, 'count': count}
            for category, count in buckets.items()
        ]
        result.sort(key=lambda item: item['count'], reverse=True)
        return Response(result)

    @action(detail=False, methods=['get'], url_path='countries')
    def countries(self, request):
        """Distinct non-empty countries within the user's scope, sorted.

        Derived from stored data, so a new card with a new country appears here
        automatically once it is saved.
        """
        values = (
            self._scoped()
            .exclude(country='')
            .values_list('country', flat=True)
            .distinct()
        )
        response = Response(sorted({(v or '').strip() for v in values if (v or '').strip()}))
        # The country list rarely changes; let the browser reuse it briefly.
        response['Cache-Control'] = 'private, max-age=300'
        return response

    @action(detail=False, methods=['get'], url_path='export-xlsx')
    def export_xlsx(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = 'جهات الاتصال'
        ws.sheet_view.rightToLeft = True

        headers = [
            ('#', 6),
            ('اسم الشخص', 28),
            ('المنصب (عربي)', 22),
            ('المنصب (إنجليزي)', 22),
            ('الشركة', 28),
            ('الموبايلات', 22),
            ('الإيميلات', 28),
            ('الموقع', 28),
            ('العنوان', 30),
            ('الدولة', 16),
            ('نشاط الشركة', 35),
            ('نوع الاستثمار', 35),
            ('تفاصيل الاستثمار', 30),
            ('يحتاج مراجعة', 14),
            ('تاريخ الإضافة', 18),
        ]

        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='AAAAAA'),
            right=Side(style='thin', color='AAAAAA'),
            top=Side(style='thin', color='AAAAAA'),
            bottom=Side(style='thin', color='AAAAAA'),
        )

        for col_idx, (header_text, col_width) in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header_text)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = col_width

        ws.row_dimensions[1].height = 30
        ws.freeze_panes = 'A2'

        row_fill_even = PatternFill(start_color='EBF3FB', end_color='EBF3FB', fill_type='solid')
        data_font = Font(name='Arial', size=10)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        right_align = Alignment(horizontal='right', vertical='center', wrap_text=True)

        for row_idx, card in enumerate(self.get_queryset(), start=2):
            row_fill = row_fill_even if row_idx % 2 == 0 else None
            row_data = [
                card.sequence_number,
                card.person_name,
                card.job_title_ar or '',
                card.job_title_en or '',
                card.company_name,
                ' | '.join(card.mobile_numbers or []),
                ' | '.join(card.emails or []),
                card.website,
                card.address,
                card.country,
                card.company_activity,
                card.investment_type,
                card.investment_type_other,
                'نعم' if card.needs_review else 'لا',
                card.created_at.strftime('%Y-%m-%d %H:%M'),
            ]
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=excel_safe(value))
                cell.font = data_font
                cell.border = thin_border
                if col_idx == 1:
                    cell.alignment = center_align
                else:
                    cell.alignment = right_align
                if row_fill:
                    cell.fill = row_fill
            ws.row_dimensions[row_idx].height = 20

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="business-cards.xlsx"'
        wb.save(response)
        return response

    @action(detail=True, methods=['get'], url_path='image/(?P<side>front|back)')
    def image(self, request, pk=None, side=None):
        """Serve a card image only to its owner (or an admin).

        Ownership is enforced by ``get_object`` walking the scoped queryset, so
        a user cannot read another user's image even with the direct URL.
        """
        card = self.get_object()  # 404 if not visible to this user
        field = card.front_image if side == 'front' else card.back_image
        if not field:
            raise Http404('لا توجد صورة.')
        try:
            handle = field.open('rb')
        except (FileNotFoundError, ValueError):
            raise Http404('تعذّر فتح الصورة.')
        response = FileResponse(handle)
        # Card images never change once uploaded; let the browser cache them
        # privately so re-opening the image modal is instant (no re-download).
        response['Cache-Control'] = 'private, max-age=86400'
        return response


@api_view(['GET'])
def health(request):
    return Response({'ok': True, 'service': 'business-card-platform'})
