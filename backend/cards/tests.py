import tempfile
from io import BytesIO

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import override_settings
from rest_framework.test import APIClient

from .models import BusinessCard
from .services.card_data import merge_missing_card_data, merge_missing_card_images, prepare_card_data
from .services.natural_search import derive_category, normalize_arabic, parse_natural_query

User = get_user_model()


def make_user(username='user', password='StrongPass!234', **extra):
    return User.objects.create_user(username=username, password=password, **extra)


def admin_client():
    """An authenticated admin API client (admins see all cards)."""
    admin = User.objects.create_user(username='legacy_admin', password='x', is_staff=True)
    client = APIClient()
    client.force_authenticate(admin)
    return client


def make_card(**overrides):
    owner = overrides.pop('owner', None)
    data = {
        'person_name': 'Test Person',
        'company_name': 'Test Co',
        'company_activity': '',
        'investment_type': '',
        'investment_type_other': '',
        'address': '',
        'emails': [],
        'mobile_numbers': [],
        'raw_text': '',
        'status': 'new',
    }
    data.update(overrides)
    prepared = prepare_card_data(data, infer_missing_investment=False)
    return BusinessCard.objects.create(owner=owner, **prepared)


class NormalizeArabicTests(TestCase):
    def test_hamza_variants_fold_together(self):
        self.assertEqual(normalize_arabic('الأمن'), normalize_arabic('الامن'))
        self.assertEqual(normalize_arabic('إسمنت'), normalize_arabic('اسمنت'))
        self.assertEqual(normalize_arabic('آلات'), normalize_arabic('الات'))

    def test_ta_marbuta_and_alef_maksura(self):
        self.assertEqual(normalize_arabic('حماة'), normalize_arabic('حماه'))
        self.assertEqual(normalize_arabic('مصطفى'), normalize_arabic('مصطفي'))

    def test_latin_lowercased(self):
        self.assertEqual(normalize_arabic('WATER'), 'water')


class ParseNaturalQueryTests(TestCase):
    def test_water_activity_arabic(self):
        parsed = parse_natural_query('عرضلي شركات المياه')
        self.assertEqual(parsed['activity_keyword'], 'مياه')
        self.assertEqual(parsed['text'], '')

    def test_electricity_english(self):
        parsed = parse_natural_query('show me electricity companies')
        self.assertEqual(parsed['activity_keyword'], 'كهرباء')

    def test_country_turkish(self):
        parsed = parse_natural_query('الشركات التركية')
        self.assertEqual(parsed['country'], 'تركيا')

    def test_city_damascus(self):
        parsed = parse_natural_query('بدي الشركات الموجودة في دمشق')
        self.assertEqual(parsed['city'], 'دمشق')

    def test_needs_review(self):
        parsed = parse_natural_query('عرض الكروت التي تحتاج مراجعة')
        self.assertEqual(parsed['status'], 'needs_review')

    def test_missing_email(self):
        parsed = parse_natural_query('عرضلي الشركات التي ليس لديها بريد الكتروني')
        self.assertTrue(parsed['missing_email'])

    def test_missing_phone_english(self):
        parsed = parse_natural_query('companies with no phone')
        self.assertTrue(parsed['missing_phone'])

    def test_plain_text_falls_back(self):
        parsed = parse_natural_query('SAMIROCK')
        self.assertIsNone(parsed['activity_keyword'])
        self.assertIsNone(parsed['city'])
        self.assertIsNone(parsed['country'])
        self.assertEqual(parsed['text'], 'SAMIROCK')

    def test_multiword_leftover_has_no_dangling_fragments(self):
        parsed = parse_natural_query('بدي شركات امن سيبراني')
        self.assertEqual(parsed['activity_keyword'], 'سيبراني')
        self.assertEqual(parsed['text'].split(), ['امن'])


class DeriveCategoryTests(TestCase):
    def test_blank_and_placeholder_map_to_unknown(self):
        self.assertEqual(derive_category(''), 'غير محدد')
        self.assertEqual(derive_category('نشاط غير محدد'), 'غير محدد')

    def test_dash_separated_takes_first_segment(self):
        self.assertEqual(
            derive_category('هندسية - شركة الإنشاءات والتخطيط المحدودة'),
            'هندسية',
        )

    def test_slash_separated_takes_first_segment(self):
        self.assertEqual(derive_category('تكنو فيست تركيا/استشارات تجارية'), 'تكنو فيست تركيا')

    def test_plain_value_used_as_is(self):
        self.assertEqual(derive_category('مقاولات'), 'مقاولات')


class NaturalSearchApiTests(TestCase):
    def setUp(self):
        self.client = admin_client()
        make_card(company_name='Water Co', company_activity='شركة مياه وصرف صحي')
        make_card(company_name='Volt Co', company_activity='كهرباء وطاقة')
        make_card(company_name='Damascus Trading', address='دمشق - سوريا')
        make_card(company_name='Istanbul Textiles', address='اسطنبول - تركيا')
        make_card(company_name='No Email Co', emails=[], status='needs_review')
        make_card(company_name='CIPER', company_activity='اتصالات - مجال الأمن السيبراني')
        make_card(company_name='Engineering Consult Co', company_activity='استشارات هندسية')

    def test_search_water_companies(self):
        response = self.client.get('/api/cards/', {'q': 'عرضلي شركات المياه'})
        self.assertEqual(response.status_code, 200)
        names = [c['company_name'] for c in response.data['results']]
        self.assertIn('Water Co', names)
        self.assertNotIn('Volt Co', names)

    def test_search_turkish_companies(self):
        response = self.client.get('/api/cards/', {'q': 'الشركات التركية'})
        self.assertEqual(response.status_code, 200)
        names = [c['company_name'] for c in response.data['results']]
        self.assertIn('Istanbul Textiles', names)

    def test_search_damascus(self):
        response = self.client.get('/api/cards/', {'q': 'في دمشق'})
        self.assertEqual(response.status_code, 200)
        names = [c['company_name'] for c in response.data['results']]
        self.assertIn('Damascus Trading', names)

    def test_search_needs_review(self):
        response = self.client.get('/api/cards/', {'q': 'تحتاج مراجعة'})
        self.assertEqual(response.status_code, 200)
        names = [c['company_name'] for c in response.data['results']]
        self.assertIn('No Email Co', names)

    def test_search_cybersecurity_multiword_and_hamza_variant(self):
        response = self.client.get('/api/cards/', {'q': 'بدي شركات امن سيبراني'})
        self.assertEqual(response.status_code, 200)
        names = [c['company_name'] for c in response.data['results']]
        self.assertIn('CIPER', names)

    def test_search_engineering_compound_phrase(self):
        response = self.client.get('/api/cards/', {'q': 'بدي شركات هندسية'})
        self.assertEqual(response.status_code, 200)
        names = [c['company_name'] for c in response.data['results']]
        self.assertIn('Engineering Consult Co', names)

    def test_stats_by_category(self):
        response = self.client.get('/api/cards/stats-by-category/')
        self.assertEqual(response.status_code, 200)
        categories = {item['category']: item['count'] for item in response.data}
        self.assertIn('شركة مياه وصرف صحي', categories)
        self.assertIn('غير محدد', categories)
        self.assertIn('اتصالات', categories)

    def test_stats_by_investment_type_groups_other_choice(self):
        make_card(company_name='Cotton One', investment_type='مؤسسة الحلج و الاقطان')
        make_card(company_name='Cotton One', investment_type='مؤسسة الحلج و الاقطان')
        make_card(company_name='Custom Investment', investment_type='غير ذلك', investment_type_other='تصنيف خاص')

        response = self.client.get('/api/cards/stats-by-category/', {'field': 'investment_type'})
        self.assertEqual(response.status_code, 200)
        categories = {item['category']: item['count'] for item in response.data}
        self.assertEqual(categories['مؤسسة الحلج و الاقطان'], 1)
        self.assertEqual(categories['غير ذلك'], 1)
        self.assertNotIn('تصنيف خاص', categories)

    def test_category_filter_matches_bucket_not_substring(self):
        response = self.client.get('/api/cards/', {'category': 'اتصالات'})
        self.assertEqual(response.status_code, 200)
        names = [c['company_name'] for c in response.data['results']]
        self.assertIn('CIPER', names)
        self.assertNotIn('Engineering Consult Co', names)


class BusinessCardUpdateApiTests(TestCase):
    def setUp(self):
        self.client = admin_client()

    def test_patch_base_job_title_replaces_existing_bilingual_title(self):
        card = make_card(job_title='Old Arabic\nOld English', job_title_ar='Old Arabic', job_title_en='Old English')
        response = self.client.patch(f'/api/cards/{card.id}', {'job_title': 'New Position'}, format='json')
        self.assertEqual(response.status_code, 200)
        card.refresh_from_db()
        self.assertEqual(card.job_title, 'New Position')
        self.assertEqual(card.job_title_ar, 'New Position')
        self.assertEqual(card.job_title_en, '')

    def test_patch_base_job_title_can_clear_existing_bilingual_title(self):
        card = make_card(job_title='Old Arabic\nOld English', job_title_ar='Old Arabic', job_title_en='Old English')
        response = self.client.patch(f'/api/cards/{card.id}', {'job_title': ''}, format='json')
        self.assertEqual(response.status_code, 200)
        card.refresh_from_db()
        self.assertEqual(card.job_title, '')
        self.assertEqual(card.job_title_ar, '')
        self.assertEqual(card.job_title_en, '')


class BusinessCardCreateDuplicateApiTests(TestCase):
    def setUp(self):
        self.client = admin_client()
        self.card_data = {
            'person_name': 'Duplicate Person',
            'company_name': 'Duplicate Co',
            'company_activity': 'شركة خدمات',
            'investment_type': 'غير ذلك',
            'investment_type_other': 'نوع خاص',
            'address': 'شارع الاختبار',
            'emails': ['dup@example.com'],
            'mobile_numbers': ['+966500000000'],
            'raw_text': 'Duplicate card',
            'status': 'new',
        }
        BusinessCard.objects.create(**prepare_card_data(self.card_data, infer_missing_investment=False))

    def test_duplicate_create_returns_200_and_existing_card(self):
        response = self.client.post('/api/cards/', self.card_data, format='json')
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.data.get('duplicate'))
        self.assertIn('existing_card', response.data)
        self.assertEqual(response.data['existing_card']['company_name'], 'Duplicate Co')
        self.assertEqual(response.data['existing_card']['emails'], ['dup@example.com'])


class DuplicateMergeTests(TestCase):
    def test_duplicate_attempt_fills_missing_existing_fields_only(self):
        card = make_card(
            person_name='Existing Person', person_name_ar='Existing Person',
            company_name='Existing Co', emails=['old@example.com'],
            mobile_numbers=[], website='', address='', company_activity='',
        )
        new_data = prepare_card_data(
            {
                'person_name': 'Different OCR Name', 'person_name_ar': 'Different OCR Name',
                'company_name': 'Existing Co', 'emails': ['old@example.com', 'new@example.com'],
                'mobile_numbers': ['+963 944 123 456'], 'website': 'example.com',
                'address': 'Damascus', 'company_activity': 'Consulting', 'confidence': 0.82,
            },
            infer_missing_investment=False,
        )
        updated_fields = merge_missing_card_data(card, new_data)
        card.refresh_from_db()
        self.assertIn('mobile_numbers', updated_fields)
        self.assertIn('website', updated_fields)
        self.assertEqual(card.person_name, 'Existing Person')
        self.assertEqual(card.emails, ['old@example.com', 'new@example.com'])
        self.assertEqual(card.website, 'https://example.com')
        self.assertEqual(card.address, 'Damascus')
        self.assertEqual(card.company_activity, 'Consulting')
        self.assertEqual(card.confidence, 0.82)

    def test_duplicate_merge_does_not_replace_base_name_when_language_fields_are_blank(self):
        card = make_card(person_name='Existing Person', company_name='Existing Co', emails=['old@example.com'])
        card.person_name_ar = ''
        card.person_name_en = ''
        card.save(update_fields=['person_name_ar', 'person_name_en'])
        new_data = prepare_card_data(
            {
                'person_name': 'Different OCR Name', 'company_name': 'Existing Co',
                'emails': ['old@example.com'], 'website': 'example.com',
            },
            infer_missing_investment=False,
        )
        updated_fields = merge_missing_card_data(card, new_data)
        card.refresh_from_db()
        self.assertIn('website', updated_fields)
        self.assertEqual(card.person_name, 'Existing Person')
        self.assertEqual(card.person_name_ar, '')
        self.assertEqual(card.person_name_en, '')

    @override_settings(MEDIA_ROOT=tempfile.gettempdir())
    def test_duplicate_merge_fills_missing_images_without_replacing_existing_images(self):
        card = make_card(emails=['old@example.com'])
        card.front_image.save(
            'existing-front.jpg',
            SimpleUploadedFile('existing-front.jpg', b'existing-front', content_type='image/jpeg'),
            save=True,
        )
        existing_front_name = card.front_image.name
        updated_fields = merge_missing_card_images(
            card,
            SimpleUploadedFile('new-front.jpg', b'new-front', content_type='image/jpeg'),
            SimpleUploadedFile('new-back.jpg', b'new-back', content_type='image/jpeg'),
        )
        card.refresh_from_db()
        self.assertEqual(updated_fields, ['back_image'])
        self.assertEqual(card.front_image.name, existing_front_name)
        self.assertTrue(card.back_image.name.endswith('new-back.jpg'))


def _png_bytes():
    # 1x1 PNG so PIL.verify() accepts the upload.
    import base64
    return base64.b64decode(
        b'iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8z8BQDwAEhQGAhKmMIQAAAABJRU5ErkJggg=='
    )


class AuthApiTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.password = 'StrongPass!234'
        self.user = make_user('zaina', self.password, email='zaina@example.com')

    def test_login_with_username(self):
        r = self.client.post('/api/auth/login', {'username': 'zaina', 'password': self.password}, format='json')
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r.data['username'], 'zaina')

    def test_login_with_email(self):
        r = self.client.post('/api/auth/login', {'username': 'zaina@example.com', 'password': self.password}, format='json')
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r.data['username'], 'zaina')

    def test_login_wrong_password_rejected(self):
        r = self.client.post('/api/auth/login', {'username': 'zaina', 'password': 'nope'}, format='json')
        self.assertEqual(r.status_code, 400)

    def test_change_password_requires_current(self):
        self.client.force_authenticate(self.user)
        bad = self.client.post('/api/auth/change-password', {
            'current_password': 'wrong', 'new_password': 'AnotherPass!99', 'new_password_confirm': 'AnotherPass!99',
        }, format='json')
        self.assertEqual(bad.status_code, 400)
        ok = self.client.post('/api/auth/change-password', {
            'current_password': self.password, 'new_password': 'AnotherPass!99', 'new_password_confirm': 'AnotherPass!99',
        }, format='json')
        self.assertEqual(ok.status_code, 200)
        self.user.refresh_from_db()
        self.assertTrue(self.user.check_password('AnotherPass!99'))

    def test_user_cannot_escalate_to_admin_via_profile(self):
        self.client.force_authenticate(self.user)
        r = self.client.patch('/api/auth/profile', {
            'first_name': 'Z', 'is_staff': True, 'is_superuser': True,
        }, format='json')
        self.assertEqual(r.status_code, 200)
        self.user.refresh_from_db()
        self.assertFalse(self.user.is_staff)
        self.assertFalse(self.user.is_superuser)

    def test_registration_disabled_by_default(self):
        r = self.client.post('/api/auth/register', {
            'username': 'newbie', 'email': 'n@e.com', 'password': 'StrongPass!234',
            'password_confirm': 'StrongPass!234',
        }, format='json')
        self.assertEqual(r.status_code, 403)


class AdminUserApiTests(TestCase):
    def test_admin_can_create_user_without_superuser(self):
        admin = make_user('boss', is_staff=True)
        client = APIClient()
        client.force_authenticate(admin)
        r = client.post('/api/admin/users/', {
            'username': 'employee', 'email': 'emp@example.com', 'first_name': 'Emp',
            'is_active': True, 'is_staff': False, 'is_superuser': True, 'password': 'StrongPass!234',
        }, format='json')
        self.assertEqual(r.status_code, 201)
        created = User.objects.get(username='employee')
        self.assertFalse(created.is_superuser)  # never granted via API

    def test_regular_user_cannot_manage_users(self):
        user = make_user('plain')
        client = APIClient()
        client.force_authenticate(user)
        self.assertEqual(client.get('/api/admin/users/').status_code, 403)


class CardOwnershipApiTests(TestCase):
    def setUp(self):
        self.alice = make_user('alice')
        self.bob = make_user('bob')
        self.admin = make_user('root', is_staff=True)
        self.alice_card = make_card(owner=self.alice, company_name='Alice Co', emails=['alice@x.com'])
        self.bob_card = make_card(owner=self.bob, company_name='Bob Co', emails=['bob@x.com'])

    def _client(self, user):
        c = APIClient()
        c.force_authenticate(user)
        return c

    def test_user_sees_only_own_cards(self):
        r = self._client(self.alice).get('/api/cards/')
        names = [c['company_name'] for c in r.data['results']]
        self.assertEqual(names, ['Alice Co'])

    def test_admin_sees_all_cards(self):
        r = self._client(self.admin).get('/api/cards/')
        names = {c['company_name'] for c in r.data['results']}
        self.assertSetEqual(names, {'Alice Co', 'Bob Co'})

    def test_user_cannot_read_other_users_card(self):
        r = self._client(self.alice).get(f'/api/cards/{self.bob_card.id}')
        self.assertEqual(r.status_code, 404)

    def test_user_cannot_update_or_delete_other_users_card(self):
        c = self._client(self.alice)
        self.assertEqual(c.patch(f'/api/cards/{self.bob_card.id}', {'company_name': 'Hacked'}, format='json').status_code, 404)
        self.assertEqual(c.delete(f'/api/cards/{self.bob_card.id}').status_code, 404)
        self.bob_card.refresh_from_db()
        self.assertEqual(self.bob_card.company_name, 'Bob Co')

    def test_backend_sets_owner_on_create_ignoring_client_owner(self):
        c = self._client(self.alice)
        r = c.post('/api/cards/', {
            'person_name': 'New', 'company_name': 'Newly', 'emails': ['new@x.com'],
            'owner': self.bob.id, 'owner_id': self.bob.id,
        }, format='json')
        self.assertIn(r.status_code, (200, 201))
        card = BusinessCard.objects.get(company_name='Newly')
        self.assertEqual(card.owner_id, self.alice.id)

    def test_stats_are_scoped_to_user(self):
        r = self._client(self.alice).get('/api/cards/stats')
        self.assertEqual(r.data['total'], 1)

    def test_export_returns_only_own_cards(self):
        from openpyxl import load_workbook
        r = self._client(self.alice).get('/api/cards/export-xlsx')
        self.assertEqual(r.status_code, 200)
        wb = load_workbook(BytesIO(r.content))
        ws = wb.active
        companies = [row[4].value for row in ws.iter_rows(min_row=2)]
        self.assertEqual(companies, ['Alice Co'])

    def test_duplicate_detection_does_not_reveal_other_users_cards(self):
        # Bob creates a card with Alice's email; must NOT be flagged as duplicate.
        c = self._client(self.bob)
        r = c.post('/api/cards/', {
            'person_name': 'X', 'company_name': 'BobDup', 'emails': ['alice@x.com'],
        }, format='json')
        self.assertIn(r.status_code, (200, 201))
        self.assertFalse(r.data.get('duplicate', False))
        self.assertTrue(BusinessCard.objects.filter(company_name='BobDup', owner=self.bob).exists())

    @override_settings(MEDIA_ROOT=tempfile.gettempdir())
    def test_other_users_image_is_inaccessible(self):
        self.bob_card.front_image.save(
            'bob-front.png', SimpleUploadedFile('bob-front.png', _png_bytes(), content_type='image/png'), save=True,
        )
        alice = self._client(self.alice)
        self.assertEqual(alice.get(f'/api/cards/{self.bob_card.id}/image/front').status_code, 404)
        bob = self._client(self.bob)
        self.assertEqual(bob.get(f'/api/cards/{self.bob_card.id}/image/front').status_code, 200)

    def test_anonymous_cannot_list_cards(self):
        self.assertEqual(APIClient().get('/api/cards/').status_code, 403)


class CountryFilterApiTests(TestCase):
    def setUp(self):
        self.alice = make_user('c_alice')
        self.bob = make_user('c_bob')
        self.a1 = make_card(owner=self.alice, company_name='A-SY', address='سوريا - دمشق')
        self.a2 = make_card(owner=self.alice, company_name='A-TR', address='تركيا - اسطنبول')
        self.b1 = make_card(owner=self.bob, company_name='B-SA', address='السعودية - الرياض')

    def _client(self, user):
        c = APIClient(); c.force_authenticate(user); return c

    def test_country_derived_on_create(self):
        self.assertEqual(self.a1.country, 'سوريا')
        self.assertEqual(self.a2.country, 'تركيا')

    def test_countries_endpoint_is_scoped(self):
        r = self._client(self.alice).get('/api/cards/countries')
        self.assertEqual(r.status_code, 200)
        self.assertEqual(sorted(r.data), sorted(['سوريا', 'تركيا']))
        # Bob's country must not leak to Alice.
        self.assertNotIn('السعودية', r.data)

    def test_country_filter_returns_only_matching_and_owned(self):
        r = self._client(self.alice).get('/api/cards/', {'country': 'سوريا'})
        names = [c['company_name'] for c in r.data['results']]
        self.assertEqual(names, ['A-SY'])


class AdminSetPasswordAndUserCardsTests(TestCase):
    def setUp(self):
        self.admin = make_user('root2', is_staff=True)
        self.alice = make_user('c2_alice', password='OldPass!234')
        self.bob = make_user('c2_bob')
        make_card(owner=self.alice, company_name='Alice-1')
        make_card(owner=self.alice, company_name='Alice-2')
        make_card(owner=self.bob, company_name='Bob-1')

    def _c(self, user):
        c = APIClient(); c.force_authenticate(user); return c

    def test_admin_sets_user_password(self):
        r = self._c(self.admin).post(f'/api/admin/users/{self.alice.id}/set-password/',
                                     {'new_password': 'BrandNew!987'}, format='json')
        self.assertEqual(r.status_code, 200)
        self.alice.refresh_from_db()
        self.assertTrue(self.alice.check_password('BrandNew!987'))

    def test_regular_user_cannot_set_password(self):
        r = self._c(self.bob).post(f'/api/admin/users/{self.alice.id}/set-password/',
                                    {'new_password': 'x'}, format='json')
        self.assertEqual(r.status_code, 403)

    def test_admin_can_view_specific_user_cards_via_owner_filter(self):
        r = self._c(self.admin).get('/api/cards/', {'owner': self.alice.id})
        names = sorted(c['company_name'] for c in r.data['results'])
        self.assertEqual(names, ['Alice-1', 'Alice-2'])

    def test_regular_user_owner_filter_ignored(self):
        # Bob passing owner=alice must still only see his own cards.
        r = self._c(self.bob).get('/api/cards/', {'owner': self.alice.id})
        names = [c['company_name'] for c in r.data['results']]
        self.assertEqual(names, ['Bob-1'])


class MergeDuplicateCardsTests(TestCase):
    def test_merge_unions_data_without_loss(self):
        from cards.services.merge import merge_duplicate_cards
        owner = make_user('m_owner')
        c1 = make_card(owner=owner, company_name='DupCo', emails=['a@x.com'], mobile_numbers=['+963944111111'])
        base = c1.duplicate_hash
        c2 = make_card(owner=owner, company_name='DupCo', emails=['b@x.com'], mobile_numbers=['+963944222222'], website='ex.com')
        c2.duplicate_hash = base + ':salt'
        c2.save(update_fields=['duplicate_hash'])

        res = merge_duplicate_cards(BusinessCard.objects.filter(owner=owner), apply=True)
        self.assertEqual(res['duplicate_groups'], 1)
        self.assertEqual(BusinessCard.objects.filter(owner=owner).count(), 1)
        primary = BusinessCard.objects.get(owner=owner)
        self.assertEqual(sorted(primary.emails), ['a@x.com', 'b@x.com'])
        self.assertEqual(len(primary.mobile_numbers), 2)

    def test_no_merge_for_distinct_cards(self):
        from cards.services.merge import merge_duplicate_cards
        owner = make_user('m_owner2')
        make_card(owner=owner, company_name='One', emails=['one@x.com'])
        make_card(owner=owner, company_name='Two', emails=['two@x.com'])
        res = merge_duplicate_cards(BusinessCard.objects.filter(owner=owner), apply=True)
        self.assertEqual(res['duplicate_groups'], 0)
        self.assertEqual(BusinessCard.objects.filter(owner=owner).count(), 2)
